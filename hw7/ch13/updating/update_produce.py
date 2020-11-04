#! python3
# corrects costs in produce spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# dictionary of produce types and updated prices
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

for row_number in range(2, sheet.max_row):

    produce_name = sheet.cell(row=row_number, column=1).value

    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_number, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')