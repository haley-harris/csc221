#!/usr/bin/env python3
''' Mulitplication table maker

Author: Haley Harris

'''

import openpyxl, argparse
from openpyxl.styles import Font


def create_multiplication_table(N):
    '''Given a number N, create an NxN multiplication table in an
    Excel spreadsheet. This sheet's filename should be "output.xlsx"

    Args:
      N (int): size of multiplication table

    Returns:
      None

    '''

    wb = openpyxl.Workbook()

    sheet = wb.active
    # set bold font
    bold_font = Font(bold=True)

    # creates chart numbering
    for i in range(1, int(N) + 1):
        # assign values to rows in first column
        row_cell = sheet.cell(row= i + 1, column= 1)
        row_cell.value = i
        row_cell.font = bold_font
        # assign values to columns in first row
        col_cell = sheet.cell(row= 1, column= i + 1)
        col_cell.value = i
        col_cell.font = bold_font

        for j in range(1, int(N) + 1):
            sheet.cell(row = i + 1, column = j + 1).value = i * j
    
    wb.save('output.xlsx')
    

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('N', help='size of multiplication table')

    args = parser.parse_args()

    create_multiplication_table(args.N)

if __name__=='__main__':
    main()